from __future__ import annotations

import os
import uuid
from pathlib import Path
from typing import List

from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from mrz import extract_passport_info
from vin import extract_vin_info
from export_declaration import extract_export_declaration_info
from car365 import fetch_car365_vehicle_info

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
TEMPLATE_PATH = BASE_DIR / "packing_list_template.xlsx"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024

COUNTRY_OPTIONS = [
    "Kyrgyzstan",
    "Kazakhstan",
    "Uzbekistan",
    "Tajikistan",
    "Russia",
    "Mongolia",
    "Other",
]


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", country_options=COUNTRY_OPTIONS)


@app.route("/generate", methods=["POST"])
def generate_packing_list():
    container_number = (request.form.get("container_number") or "").strip()
    seal_number = (request.form.get("seal_number") or "").strip()
    page_date = (request.form.get("page_date") or "").strip()
    country = (request.form.get("country") or "").strip()
    other_country = (request.form.get("other_country") or "").strip()
    city = (request.form.get("city") or "").strip()

    if country == "Other":
        country = other_country.strip()

    try:
        vehicle_count = int(request.form.get("vehicle_count", "0"))
    except ValueError:
        flash("Invalid number of vehicles specified.")
        return redirect(url_for("index"))

    if not container_number:
        flash("Please specify the container number.")
        return redirect(url_for("index"))

    if not seal_number:
        flash("Please specify the seal number.")
        return redirect(url_for("index"))

    if not page_date:
        flash("Please specify the date for G4.")
        return redirect(url_for("index"))

    if not country:
        flash("Please select a country.")
        return redirect(url_for("index"))

    if not city:
        flash("Please specify the city.")
        return redirect(url_for("index"))

    if vehicle_count < 1 or vehicle_count > 10:
        flash("The number of vehicles must be between 1 and 10.")
        return redirect(url_for("index"))

    if not TEMPLATE_PATH.exists():
        flash("Packing list template not found.")
        return redirect(url_for("index"))

    request_id = uuid.uuid4().hex
    destination = f"{country}, {city}"
    vehicle_rows: List[dict] = []

    for i in range(1, vehicle_count + 1):
        passport_file = request.files.get(f"passport_{i}")
        vin_file = request.files.get(f"vin_{i}")
        export_file = request.files.get(f"export_{i}")

        if passport_file is None or passport_file.filename == "":
            flash(f"Passport photo not uploaded for vehicle #{i}.")
            return redirect(url_for("index"))

        if vin_file is None or vin_file.filename == "":
            flash(f"VIN sticker photo not uploaded for vehicle #{i}.")
            return redirect(url_for("index"))

        if export_file is None or export_file.filename == "":
            flash(f"Export declaration not uploaded for vehicle #{i}.")
            return redirect(url_for("index"))

        vehicle_dir = UPLOAD_DIR / request_id / f"vehicle_{i}"
        vehicle_dir.mkdir(parents=True, exist_ok=True)

        passport_path = vehicle_dir / secure_filename(passport_file.filename)
        vin_path = vehicle_dir / secure_filename(vin_file.filename)
        export_path = vehicle_dir / secure_filename(export_file.filename)

        passport_file.save(passport_path)
        vin_file.save(vin_path)
        export_file.save(export_path)

        passport_info = extract_passport_info(passport_path)
        vin_info = extract_vin_info(vin_path)
        export_info = extract_export_declaration_info(export_path)

        vin_string = (vin_info.get("vin_string") or "").strip()

        car365_info = {
            "returned_chassis_no": "",
            "vehicle_trademark": "",
            "vehicle_first_registration": "",
            "vehicle_fuel": "",
            "vehicle_displacement": "",
            "status": "SKIPPED",
            "error": "VIN is missing",
        }

        if vin_string:
            car365_info = fetch_car365_vehicle_info(vin_string)

        surname = (passport_info.get("surname") or "").strip()
        given_name = (passport_info.get("given_name") or "").strip()
        full_name = " ".join(part for part in [surname, given_name] if part).strip()

        vehicle_rows.append(
            {
                "vehicle_no": i,
                "kg_number": f"KG-{i:02d}",
                "page_date": page_date,
                "container_number": container_number,
                "seal_number": seal_number,
                "destination": destination,
                "full_name": full_name,
                "document_id": (passport_info.get("document_id") or "").strip(),
                "vin_string": vin_string,
                "hs_code": (export_info.get("hs_code") or "").strip(),
                "weight": (export_info.get("weight") or "").strip(),
                "vehicle_trademark": (car365_info.get("vehicle_trademark") or "").strip(),
                "vehicle_first_registration": (car365_info.get("vehicle_first_registration") or "").strip(),
                "vehicle_fuel": (car365_info.get("vehicle_fuel") or "").strip(),
                "vehicle_displacement": (car365_info.get("vehicle_displacement") or "").strip(),
                "returned_chassis_no": (car365_info.get("returned_chassis_no") or "").strip(),
                "passport_status": passport_info.get("status", ""),
                "passport_error": passport_info.get("error", ""),
                "vin_status": vin_info.get("status", ""),
                "vin_error": vin_info.get("error", ""),
                "export_status": export_info.get("status", ""),
                "export_error": export_info.get("error", ""),
                "car365_status": car365_info.get("status", ""),
                "car365_error": car365_info.get("error", ""),
            }
        )

    output_filename = f"packing_list_{safe_filename_part(container_number)}_{request_id[:8]}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    build_packing_list_workbook(output_path, vehicle_rows)

    return send_file(
        output_path,
        as_attachment=True,
        download_name=output_filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def safe_filename_part(value: str) -> str:
    value = (value or "").strip()
    allowed = "".join(ch for ch in value if ch.isalnum() or ch in ("-", "_"))
    return allowed or "packing_list"


def build_packing_list_workbook(output_path: Path, rows: List[dict]) -> None:
    wb = load_workbook(TEMPLATE_PATH)
    template_sheet = wb[wb.sheetnames[0]]

    while len(wb.sheetnames) < len(rows):
        wb.copy_worksheet(template_sheet)

    while len(wb.sheetnames) > len(rows):
        wb.remove(wb[wb.sheetnames[-1]])

    for sheet_name, item in zip(wb.sheetnames, rows):
        ws = wb[sheet_name]
        fill_packing_sheet(ws, item)
        try:
            ws.title = item["kg_number"]
        except ValueError:
            pass

    wb.save(output_path)


def fill_packing_sheet(ws, item: dict) -> None:
    ws["B8"] = item.get("full_name", "")
    ws["B9"] = item.get("document_id", "")
    ws["B29"] = item.get("vin_string", "")
    ws["E16"] = item.get("container_number", "")
    ws["E17"] = item.get("seal_number", "")
    ws["G3"] = item.get("kg_number", "")
    ws["G4"] = item.get("page_date", "")
    ws["C17"] = item.get("destination", "")

    ws["A25"] = item.get("hs_code", "")

    weight_value = item.get("weight", "")
    if weight_value:
        try:
            ws["G25"] = float(weight_value)
        except ValueError:
            ws["G25"] = weight_value

    ws["B25"] = item.get("vehicle_trademark", "")
    ws["B26"] = item.get("vehicle_fuel", "")
    ws["B27"] = item.get("vehicle_first_registration", "")
    ws["B32"] = item.get("vehicle_displacement", "")

    if item.get("passport_status") == "ERROR" and item.get("passport_error"):
        ws["B12"] = f"Passport error: {item['passport_error']}"

    if item.get("vin_status") == "ERROR" and item.get("vin_error"):
        ws["B31"] = f"VIN error: {item['vin_error']}"

    if item.get("export_status") == "ERROR" and item.get("export_error"):
        ws["A27"] = f"Export error: {item['export_error']}"

    if item.get("car365_status") not in ("", "OK", "SKIPPED") and item.get("car365_error"):
        ws["B34"] = f"Car365 error: {item['car365_error']}"


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)